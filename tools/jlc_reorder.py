#!/usr/bin/env python3
"""
Reorder aluminum extrusion processing codes to JLC (嘉立创) standard order.

Usage:
    # Single model
    python3 tools/jlc_reorder.py TXCJ-H7-2020D-L300-LC-Z5-A10-RC-Z5-A60-DA-SC

    # Multiple models
    python3 tools/jlc_reorder.py MODEL1 MODEL2 MODEL3

    # From text file (one model per line, # for comments)
    python3 tools/jlc_reorder.py -f bom_list.txt

    # Excel mode: convert JLC型号 column and output new xlsx
    python3 tools/jlc_reorder.py -x input.xlsx -o output.xlsx

    # Excel mode: specify column name (default: JLC型号)
    python3 tools/jlc_reorder.py -x input.xlsx -o output.xlsx -c "我的型号"

    # Interactive mode (stdin, Ctrl+D to exit)
    python3 tools/jlc_reorder.py

    # Pipe mode
    echo "TXCJ-H7-2020D-L260-LC-Z5-A36-DA-SC" | python3 tools/jlc_reorder.py

JLC processing code order:
    1.  SC          铣端面
    2.  LA          左端端面攻牙
    3.  RA          右端端面攻牙
    4.  DA          两端端面攻牙
    5.  LU/RU/...   45°斜切
    6.  LH/RH/...   盲孔
    7.  LC          左端水平沉头孔
    8.  RC          右端水平沉头孔
    9.  LE          左端垂直沉头孔
    10. RE          右端垂直沉头孔
    11. LBC/RBC     背面水平沉头
    12. LBE/RBE     背面垂直沉头
    13. LK/RK       水平扳手孔
    14. LM/RM       垂直扳手孔
"""

import re
import sys
from pathlib import Path

JLC_ORDER = {
    'SC': 0,
    'LA': 1, 'RA': 2, 'DA': 3,
    'LU': 4, 'RU': 4, 'LD': 4, 'RD': 4,
    'LF': 4, 'RF': 4, 'LB': 4, 'RB': 4,
    'LH': 5, 'RH': 5, 'LDH': 5, 'RDH': 5,
    'LG': 5, 'RG': 5, 'LDG': 5, 'RDG': 5,
    'LC': 6, 'RC': 7,
    'LE': 8, 'RE': 9,
    'LBC': 10, 'RBC': 10,
    'LBE': 11, 'RBE': 11,
    'LK': 12, 'RK': 12,
    'LM': 13, 'RM': 13,
}

PROCESSING_CODES = set(JLC_ORDER.keys())

PARAM_RE = re.compile(r'^(?:SD|[ZABCDE])\d+$')
LENGTH_RE = re.compile(r'^L\d+$')
TXCJ_RE = re.compile(r'TXCJ.*L\d{2,4}(-|$)')


def is_param(token):
    return bool(PARAM_RE.match(token))


def is_length(token):
    return bool(LENGTH_RE.match(token))


def is_extrusion_model(value):
    """Check if a string looks like a TXCJ extrusion model number."""
    return bool(TXCJ_RE.match(str(value)))


def parse_model(model_str):
    """Parse model string into (base_parts, [(code, [params]), ...])."""
    parts = model_str.strip().split('-')

    base_end = 0
    for i, part in enumerate(parts):
        if is_length(part):
            base_end = i + 1
            break

    base = parts[:base_end]
    remaining = parts[base_end:]

    codes = []
    current_code = None
    current_params = []

    for token in remaining:
        if token in PROCESSING_CODES:
            if current_code is not None:
                codes.append((current_code, current_params))
            current_code = token
            current_params = []
        elif is_param(token):
            current_params.append(token)
        else:
            current_params.append(token)

    if current_code is not None:
        codes.append((current_code, current_params))

    return base, codes


def reconstruct(base, codes):
    """Reconstruct model string from base and sorted codes."""
    parts = list(base)
    for code, params in codes:
        parts.append(code)
        parts.extend(params)
    return '-'.join(parts)


def convert(model_str):
    """Convert a model number to JLC standard order."""
    if not model_str or not is_extrusion_model(model_str):
        return model_str
    base, codes = parse_model(model_str)
    sorted_codes = sorted(codes, key=lambda item: JLC_ORDER.get(item[0], 99))
    return reconstruct(base, sorted_codes)


def process_line(line):
    """Process a single input line, return (input, output) or None."""
    line = line.strip()
    if not line or line.startswith('#'):
        return None
    return line, convert(line)


def find_column_index(ws, col_name):
    """Find column index by header name (searches first 3 rows)."""
    for row in ws.iter_rows(min_row=1, max_row=3, values_only=False):
        for cell in row:
            if cell.value and col_name in str(cell.value):
                return cell.column
    return None


def process_excel(input_path, output_path, col_name):
    """Read xlsx, convert target column, write new xlsx."""
    import openpyxl
    from copy import copy

    wb = openpyxl.load_workbook(input_path)
    changed = 0
    unchanged = 0
    skipped = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        col_idx = find_column_index(ws, col_name)

        if col_idx is None:
            print(f"  Sheet '{sheet_name}': column '{col_name}' not found, skipped")
            skipped += 1
            continue

        header = ws.cell(row=1, column=col_idx).value or ws.cell(row=2, column=col_idx).value or ''
        print(f"  Sheet '{sheet_name}': found column '{header}' at col {col_idx}")

        # Find data start row (skip header rows, typically row 3 in JLC template)
        data_start = 3
        for row in range(1, 4):
            val = ws.cell(row=row, column=col_idx).value
            if val and 'JLC' in str(val):
                data_start = row + 1
                break
            if val and '型号' in str(val):
                data_start = row + 1
                break

        for row in range(data_start, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            old_val = cell.value
            if not old_val:
                continue
            old_str = str(old_val).strip()
            if not is_extrusion_model(old_str):
                continue
            new_str = convert(old_str)
            if new_str != old_str:
                cell.value = new_str
                print(f"    Row {row}: {old_str}")
                print(f"         -> {new_str}")
                changed += 1
            else:
                unchanged += 1

    wb.save(output_path)
    print(f"\n  Saved: {output_path}")
    print(f"  Changed: {changed}, Unchanged: {unchanged}, Sheets skipped: {skipped}")


def parse_args():
    args = sys.argv[1:]
    opts = {
        'excel': None,
        'output': None,
        'column': 'JLC型号',
        'models': [],
    }

    i = 0
    while i < len(args):
        if args[i] == '-x' and i + 1 < len(args):
            opts['excel'] = args[i + 1]
            i += 2
        elif args[i] == '-o' and i + 1 < len(args):
            opts['output'] = args[i + 1]
            i += 2
        elif args[i] == '-c' and i + 1 < len(args):
            opts['column'] = args[i + 1]
            i += 2
        elif args[i] == '-f' and i + 1 < len(args):
            # Text file mode (legacy)
            opts['text_file'] = args[i + 1]
            i += 2
        elif not args[i].startswith('-'):
            opts['models'].append(args[i])
            i += 1
        else:
            i += 1

    return opts


def main():
    opts = parse_args()

    # Excel mode
    if opts['excel']:
        input_path = Path(opts['excel'])
        if not input_path.exists():
            print(f"Error: file not found: {input_path}", file=sys.stderr)
            sys.exit(1)

        if opts['output']:
            output_path = Path(opts['output'])
        else:
            output_path = input_path.with_stem(input_path.stem + '_jlc')

        print(f"Processing: {input_path}")
        print(f"Target column: '{opts['column']}'")
        process_excel(str(input_path), str(output_path), opts['column'])
        return

    # Text file mode
    if 'text_file' in opts:
        with open(opts['text_file']) as f:
            for line in f:
                result = process_line(line)
                if result:
                    _, out = result
                    print(out)
        return

    # CLI args mode
    if opts['models']:
        for model in opts['models']:
            result = process_line(model)
            if result:
                inp, out = result
                print(f"  input:  {inp}")
                print(f"  output: {out}")
        return

    # Stdin / interactive mode
    if sys.stdin.isatty():
        print("Enter model numbers (one per line, Ctrl+D to exit):")

    for line in sys.stdin:
        result = process_line(line)
        if result:
            inp, out = result
            if sys.stdin.isatty():
                print(f"  input:   {inp}")
                print(f"  output:  {out}")
                print()
            else:
                print(out)


if __name__ == '__main__':
    main()
